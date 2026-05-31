from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import make_response
import io

def export_to_excel(headers, data, filename):
    """
    通用 Excel 导出工具函数
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "导出数据"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="428BCA", end_color="428BCA", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # 写入表头
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 写入数据
    for row in data:
        ws.append(row)

    # 列宽自适应
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回响应
    response = make_response(output.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    
    return response
